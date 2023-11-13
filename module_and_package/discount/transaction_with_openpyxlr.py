from openpyxl import Workbook, load_workbook

reduction = 10
file_name = "transactions.xlsx"

wb = load_workbook(file_name)
ws = wb.active

for i in range(2, ws.max_row + 1):
    price = ws[f'C{i}'].value
    price = price.replace('*€*', '').replace(',', '.')
    price = float(price) * (1 - (reduction/100))
    price = format(price, ".2f")

    ws[f'D{i}'].value = price

wb.save("New_transaction_openpyxlr.xlsx")
